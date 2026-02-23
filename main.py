from PyQt5 import QtCore, QtGui, QtWidgets
from PyQt5.QtGui import QPixmap
from PyQt5.QtCore import Qt
from PyQt5.QtWidgets import QMessageBox
from PyQt5.QtGui import QFont
import chardet
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from traceback import format_exception
import sys
from PyQt5.QtGui import QMovie

import os
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, Border, Side
from openpyxl.styles import Alignment

import datetime


class Ui_MainWindow(object):
    filePath = ""
    kvaziCounter = 0
    sheet = 0
    pathToSave = ""

    def excepthook(self, type, value, traceback):
        error_message = ''.join(format_exception(type, value, traceback))
        print(error_message)
        self.gifLabel1.hide()
        self.gifLabel2.hide()
        self.ErrorLabel.setText(error_message)
        self.Error2Label.setText(error_message)

    def detect_encoding(self, file_path):
        with open(file_path, 'rb') as file:
            raw_data = file.read()
            result = chardet.detect(raw_data)
        return result['encoding']

    def setupUi(self, MainWindow):
        sys.excepthook = self.excepthook

        MainWindow.setObjectName("MainWindow")
        MainWindow.setFixedSize(332, 249)
        icon_path = os.path.join(self.getIconPath(), "myicon.ico")
        MainWindow.setWindowIcon(QtGui.QIcon(icon_path))
        font = QtGui.QFont()
        font.setPointSize(8)
        MainWindow.setFont(font)
        self.centralwidget = QtWidgets.QWidget(MainWindow)
        self.centralwidget.setObjectName("centralwidget")
        self.tabWidget = QtWidgets.QTabWidget(self.centralwidget)
        self.tabWidget.setGeometry(QtCore.QRect(0, 0, 321, 231))
        font = QtGui.QFont()
        font.setPointSize(8)
        self.tabWidget.setFont(font)
        self.tabWidget.setObjectName("tabWidget")

        self.TenzoTab = QtWidgets.QWidget()
        self.TenzoTab.setObjectName("TenzoTab")

        self.gifLabel1 = QtWidgets.QLabel(self.TenzoTab)
        self.gifLabel1.setGeometry(130, 120, 60, 60)
        self.movie = QMovie('dinosaur.gif')
        self.movie.setScaledSize(self.gifLabel1.size())
        self.gifLabel1.setMovie(self.movie)
        self.movie.start()

        self.ButtonSave = QtWidgets.QPushButton(self.TenzoTab)
        self.ButtonSave.setGeometry(QtCore.QRect(240, 180, 75, 23))
        self.ButtonSave.setObjectName("ButtonSave")
        self.ButtonSave.clicked.connect(self.buttonPressed)

        self.ButtonGetBD = QtWidgets.QPushButton(self.TenzoTab)
        self.ButtonGetBD.setGeometry(QtCore.QRect(10, 180, 75, 23))
        self.ButtonGetBD.setObjectName("ButtonGetBD")
        self.ButtonGetBD.clicked.connect(self.buttonPressedGetBD)

        self.ButtonExpressAnal = QtWidgets.QPushButton(self.TenzoTab)
        self.ButtonExpressAnal.setGeometry(QtCore.QRect(100, 180, 125, 23))
        self.ButtonExpressAnal.setObjectName("ButtonExpressAnal")
        self.ButtonExpressAnal.clicked.connect(self.buttonPressedExpressAnal)
        fontButton = QFont()
        fontButton.setBold(True)
        self.ButtonExpressAnal.setFont(fontButton)

        self.CheckBoxToKvazi = QtWidgets.QCheckBox(self.TenzoTab)
        self.CheckBoxToKvazi.setGeometry(QtCore.QRect(10, 80, 101, 17))
        self.CheckBoxToKvazi.setObjectName("CheckBoxToKvazi")
        self.CheckBoxToKvazi.clicked.connect(self.kvaziClicked)

        self.CheckBoxToMax = QtWidgets.QCheckBox(self.TenzoTab)
        self.CheckBoxToMax.setGeometry(QtCore.QRect(10, 100, 101, 17))
        self.CheckBoxToMax.setObjectName("CheckBoxToMax")

        self.ButtonGetFilePath = QtWidgets.QToolButton(self.TenzoTab)
        self.ButtonGetFilePath.setGeometry(QtCore.QRect(280, 50, 25, 21))
        self.ButtonGetFilePath.setObjectName("ButtonGetFilePath")
        self.ButtonGetFilePath.clicked.connect(self.showFileDialog)

        self.DisplayPath = QtWidgets.QLineEdit(self.TenzoTab)
        self.DisplayPath.setGeometry(QtCore.QRect(10, 50, 261, 21))
        self.DisplayPath.setEchoMode(QtWidgets.QLineEdit.Normal)
        self.DisplayPath.setReadOnly(True)
        self.DisplayPath.setObjectName("DisplayPath")

        self.HeaderLabel = QtWidgets.QLabel(self.TenzoTab)
        self.HeaderLabel.setGeometry(QtCore.QRect(10, 10, 301, 31))
        self.HeaderLabel.setObjectName("HeaderLabel")
        icon_path = os.path.join(self.getIconPath(), "header.png")
        pixmap = QPixmap(icon_path)
        self.HeaderLabel.setPixmap(pixmap)
        self.HeaderLabel.setScaledContents(True)

        self.ErrorLabel = QtWidgets.QLabel(self.TenzoTab)
        self.ErrorLabel.setGeometry(QtCore.QRect(10, 150, 301, 31))
        self.ErrorLabel.setObjectName("ErrorLabel")
        self.ErrorLabel.setText("")
        self.ErrorLabel.setAlignment(QtCore.Qt.AlignCenter | QtCore.Qt.AlignVCenter)

        self.Button = QtWidgets.QPushButton(self.TenzoTab)
        self.Button.setGeometry(QtCore.QRect(160, 180, 75, 23))
        self.Button.setObjectName("Button")
        self.Button.hide()
        self.Button1 = QtWidgets.QPushButton(self.TenzoTab)
        self.Button1.setGeometry(QtCore.QRect(80, 180, 75, 23))
        self.Button1.setObjectName("Button1")
        self.Button1.hide()

        self.tabWidget.addTab(self.TenzoTab, "")

        self.VybroTab = QtWidgets.QWidget()
        self.VybroTab.setObjectName("VybroTab")
        self.tabWidget.addTab(self.VybroTab, "")

        self.AboutTab = QtWidgets.QWidget()
        self.AboutTab.setObjectName("AboutTab")
        self.tabWidget.addTab(self.AboutTab, "")

        self.Header3Label = QtWidgets.QLabel(self.AboutTab)
        self.Header3Label.setGeometry(QtCore.QRect(10, 10, 301, 31))
        self.Header3Label.setObjectName("HeaderLabel")
        self.Header3Label.setPixmap(pixmap)
        self.Header3Label.setScaledContents(True)

        self.InfoTextEdit = QtWidgets.QTextEdit(self.AboutTab)
        self.InfoTextEdit.setGeometry(QtCore.QRect(10, 50, 301, 150))
        self.InfoTextEdit.setReadOnly(True)
        self.InfoTextEdit.setFrameStyle(QtWidgets.QFrame.NoFrame)
        self.InfoTextEdit.setStyleSheet("background-color: transparent;")
        self.InfoTextEdit.setObjectName("InfoTextEdit")

        self.Header2Label = QtWidgets.QLabel(self.VybroTab)
        self.Header2Label.setGeometry(QtCore.QRect(10, 10, 301, 31))
        self.Header2Label.setObjectName("HeaderLabel")
        self.Header2Label.setPixmap(pixmap)
        self.Header2Label.setScaledContents(True)

        self.Button2GetFilePath = QtWidgets.QToolButton(self.VybroTab)
        self.Button2GetFilePath.setGeometry(QtCore.QRect(280, 50, 25, 21))
        self.Button2GetFilePath.setObjectName("ButtonGetFilePath")
        self.Button2GetFilePath.clicked.connect(self.showFileDialog)

        self.Display2Path = QtWidgets.QLineEdit(self.VybroTab)
        self.Display2Path.setGeometry(QtCore.QRect(10, 50, 261, 21))
        self.Display2Path.setEchoMode(QtWidgets.QLineEdit.Normal)
        self.Display2Path.setReadOnly(True)
        self.Display2Path.setObjectName("DisplayPath")

        self.Button2Save = QtWidgets.QPushButton(self.VybroTab)
        self.Button2Save.setGeometry(QtCore.QRect(240, 180, 75, 23))
        self.Button2Save.setObjectName("ButtonSave")
        self.Button2Save.clicked.connect(self.buttonPressedVybro)

        self.Error2Label = QtWidgets.QLabel(self.VybroTab)
        self.Error2Label.setGeometry(QtCore.QRect(10, 150, 301, 31))
        self.Error2Label.setObjectName("ErrorLabel")
        self.Error2Label.setText("")
        self.Error2Label.setAlignment(QtCore.Qt.AlignCenter | QtCore.Qt.AlignVCenter)

        self.gifLabel2 = QtWidgets.QLabel(self.VybroTab)
        self.gifLabel2.setGeometry(130, 120, 60, 60)
        self.movie = QMovie('dinosaur.gif')
        self.movie.setScaledSize(self.gifLabel2.size())
        self.gifLabel2.setMovie(self.movie)
        self.movie.start()

        MainWindow.setCentralWidget(self.centralwidget)
        self.statusbar = QtWidgets.QStatusBar(MainWindow)
        self.statusbar.setObjectName("statusbar")
        MainWindow.setStatusBar(self.statusbar)

        self.retranslateUi(MainWindow)
        self.tabWidget.setCurrentIndex(0)
        QtCore.QMetaObject.connectSlotsByName(MainWindow)

    def showFileDialog(self):
        options = QtWidgets.QFileDialog.Options()
        self.filePath, _ = QtWidgets.QFileDialog.getOpenFileName(MainWindow, "Выберите файл", "", "All Files (*)",
                                                                 options=options)
        self.gifLabel1.show()
        self.gifLabel2.show()

        if ".txt" in self.filePath:
            self.ErrorLabel.setText("")
            self.DisplayPath.setText(self.filePath)
            self.Display2Path.setText(self.filePath)
        else:
            self.DisplayPath.setText("Неверный формат файла")

        current_tab_index = self.tabWidget.currentIndex()
        if current_tab_index == 0:
            self.Display2Path.setText("")
        else:
            self.DisplayPath.setText("")

    def retranslateUi(self, MainWindow):
        _translate = QtCore.QCoreApplication.translate
        MainWindow.setWindowTitle(_translate("MainWindow", "Basement"))
        self.ButtonSave.setText(_translate("MainWindow", "Сохранить"))
        self.Button2Save.setText(_translate("MainWindow", "Сохранить"))

        self.ButtonGetBD.setText(_translate("MainWindow", "Создать БД"))
        self.ButtonExpressAnal.setText(_translate("MainWindow", "Экспресс анализ"))

        self.CheckBoxToKvazi.setText(_translate("MainWindow", "Квазистатика"))
        self.CheckBoxToMax.setText(_translate("MainWindow", "Максимумы"))

        self.ButtonGetFilePath.setText(_translate("MainWindow", "..."))
        self.Button2GetFilePath.setText(_translate("MainWindow", "..."))

        self.tabWidget.setTabText(self.tabWidget.indexOf(self.TenzoTab), _translate("MainWindow", "Тензометрия"))
        self.tabWidget.setTabText(self.tabWidget.indexOf(self.VybroTab), _translate("MainWindow", "Полосовой анализ"))
        self.tabWidget.setTabText(self.tabWidget.indexOf(self.AboutTab), _translate("MainWindow", "Инфо"))

        self.InfoTextEdit.setText(_translate("MainWindow",
            "Предложения и ошибки отправлять на почту:\n"
            "olga.titkovaa.a@gmail.com\n\n"
            "Поддержать разработчика:\n"
            "По номеру телефона (СБП):\n"
            "+7 (916) 426-33-29 (Т Банк)"))

    def fillCell(self, value, row, column):
        double = Side(border_style="thin", color="000000")
        border_style = Border(left=double, right=double, top=double, bottom=double)

        font_style = Font(name="Times New Roman", size=12)
        center_alignment = Alignment(horizontal='center', vertical='center')
        try:
            float(value.replace(u'\xa0', ""))
            self.sheet.cell(row=row, column=column, value=float(value.replace(u'\xa0', ""))).font = font_style
            self.sheet.cell(row=row, column=column).alignment = center_alignment
            self.sheet.cell(row=row, column=column).border = border_style
        except:
            self.sheet.cell(row=row, column=column, value=value).font = font_style
            self.sheet.cell(row=row, column=column).alignment = center_alignment
            self.sheet.cell(row=row, column=column).border = border_style

    def fiillingLeftColumn(self, numberOfLines, startRow, startColumn):
        for idx in range(1, numberOfLines):
            double = Side(border_style="thin", color="000000")
            border_style = Border(left=double, right=double, top=double, bottom=double)
            self.sheet.cell(row=startRow + idx, column=startColumn).border = border_style
            self.sheet.cell(row=startRow + idx, column=startColumn + 1).border = border_style
            self.sheet.cell(row=startRow + idx, column=startColumn + 2).border = border_style
            self.sheet.cell(row=startRow + idx, column=startColumn + 3).border = border_style
            self.sheet.cell(row=startRow + idx, column=startColumn + 4).border = border_style

    def workWithMax(self, numberOfLines, startOffset, columnOffset, checkToKZ, maximumOfColumn1, maximumOfColumn2, maximumOfColumn3, maximumOfColumnKZ):
        sdvigY = 1

        targetMax1 = maximumOfColumn1
        targetMax2 = maximumOfColumn2
        targetMax3 = maximumOfColumn3
        targetMaxKZ = maximumOfColumnKZ

        if checkToKZ:
            targetMax1 = maximumOfColumnKZ
            targetMax2 = maximumOfColumn1
            targetMax3 = maximumOfColumn2
            targetMaxKZ = maximumOfColumn3

        for row_num in range(2, numberOfLines + 2):
            value1 = self.sheet.cell(row=row_num, column=startOffset + columnOffset + 1).value
            value2 = self.sheet.cell(row=row_num, column=startOffset + columnOffset + 2).value
            value3 = self.sheet.cell(row=row_num, column=startOffset + columnOffset + 3).value
            valueKZ = self.sheet.cell(row=row_num,
                                      column=startOffset + columnOffset + 4).value if checkToKZ else None

            if value1 == targetMax1:
                self.sheet.cell(row=row_num,
                                column=startOffset + columnOffset + 1).font = openpyxl.styles.Font(
                    bold=True, name="Times New Roman", size=12)

            if value2 == targetMax2:
                self.sheet.cell(row=row_num,
                                column=startOffset + columnOffset + 2).font = openpyxl.styles.Font(
                    bold=True, name="Times New Roman", size=12)

            if value3 == targetMax3:
                self.sheet.cell(row=row_num,
                                column=startOffset + columnOffset + 3).font = openpyxl.styles.Font(
                    bold=True, name="Times New Roman", size=12)

            if checkToKZ and valueKZ == targetMaxKZ:
                self.sheet.cell(row=row_num,
                                column=startOffset + columnOffset + 4).font = openpyxl.styles.Font(
                    bold=True, name="Times New Roman", size=12)

        self.fillCell(str(targetMax1), numberOfLines + 1 + sdvigY, startOffset + columnOffset + 1)
        self.fillCell(str(targetMax2), numberOfLines + 1 + sdvigY, startOffset + columnOffset + 2)
        self.fillCell(str(targetMax3), numberOfLines + 1 + sdvigY, startOffset + columnOffset + 3)

        if checkToKZ:
            self.fillCell(str(targetMaxKZ), numberOfLines + 1 + sdvigY, startOffset + columnOffset + 4)

    def makingsheetWithRules(self, path, filenameToSave=""):

        checkToKZ = self.CheckBoxToKvazi.isChecked()
        checkToSI = True  # Check to place SI
        checkToMAX = self.CheckBoxToMax.isChecked()
        BLOCK_SIZE = 3 if checkToKZ == 0 else 4
        sdvigX = 4
        sdvigY = 1
        lowerSdvigX = 2
        sdvigToKZ = 1 if checkToKZ else 0
        workbook = openpyxl.Workbook()
        self.sheet = workbook.active
        startOffset = 2
        flagToDrawStart = True
        flagToWriteNames = False

        maximumOfColumn1 = 0; maximumOfColumn2 = 0; maximumOfColumn3 = 0; maximumOfColumnKZ = 0

        start_time = datetime.datetime.now()

        with open(path, 'r', encoding='cp1251') as file:
            lines = file.readlines()

        currentRow = 1
        numberOfLines = 0
        columnOffset = -BLOCK_SIZE

        flag = False
        index = 0

        self.fillCell("Ид-р реж.", 1 + sdvigY, 1 + lowerSdvigX)
        self.fillCell("Наименование режима", 1 + sdvigY, 2 + lowerSdvigX)
        self.fillCell("Vпр, км/ч", 1 + sdvigY, 3 + lowerSdvigX)
        self.fillCell("Hабс, м", 1 + sdvigY, 4 + lowerSdvigX)

        self.sheet.merge_cells(start_row=1 + sdvigY, end_row=3 + sdvigY, start_column=lowerSdvigX,
                               end_column=lowerSdvigX)
        self.sheet.merge_cells(start_row=1 + sdvigY, end_row=3 + sdvigY, start_column=1 + lowerSdvigX,
                          end_column=1 + lowerSdvigX)
        self.sheet.merge_cells(start_row=1 + sdvigY, end_row=3 + sdvigY, start_column=2 + lowerSdvigX,
                          end_column=2 + lowerSdvigX)
        self.sheet.merge_cells(start_row=1 + sdvigY, end_row=3 + sdvigY, start_column=3 + lowerSdvigX,
                               end_column=3 + lowerSdvigX)
        self.sheet.merge_cells(start_row=1 + sdvigY, end_row=3 + sdvigY, start_column=4 + lowerSdvigX,
                               end_column=4 + lowerSdvigX)

        self.sheet.column_dimensions["C"].width = len("Ид-р реж.") + 4
        self.sheet.column_dimensions["D"].width = len("Наименование режима") + 5
        self.sheet.column_dimensions["E"].width = len("Vпр, км/ч") + 4
        self.sheet.column_dimensions["F"].width = len("Hабс, м") + 4

        while index < len(lines):

            currentLine = lines[index].strip()
            splittedLine = currentLine.split("	")

            if "Результаты расчета" in currentLine:
                flag = True
                index += 1
                continue

            if "Параметр:" in currentLine:
                maximumOfColumn1 = 0; maximumOfColumn2 = 0; maximumOfColumn3 = 0; maximumOfColumnKZ = 0
                columnOffset += BLOCK_SIZE

                # To draw starting 2 columns
                if columnOffset >= BLOCK_SIZE and flagToDrawStart == True:
                    flagToDrawStart = False

                currentRow = 1
                if checkToSI:
                    self.fillCell(splittedLine[0][10:] + ", " + splittedLine[3][14:], currentRow + sdvigY, startOffset + columnOffset + 1 + sdvigX)
                else:
                    self.fillCell(splittedLine[0][10:], currentRow + sdvigY, startOffset + columnOffset + 1 + sdvigX)
                self.sheet.merge_cells(start_row=1 + sdvigY, end_row=1 + sdvigY, start_column=startOffset + columnOffset + 1 + sdvigX,
                                  end_column=startOffset + columnOffset + BLOCK_SIZE + sdvigX)
                index += 1
                continue

            if flag:
                if "" == splittedLine[0]:
                    if columnOffset == 0 & currentRow != 1:
                        numberOfLines = currentRow + 1
                    if checkToMAX:
                        if numberOfLines != 0:
                            self.workWithMax(numberOfLines, startOffset, columnOffset + sdvigX, checkToKZ, maximumOfColumn1,
                                        maximumOfColumn2, maximumOfColumn3, maximumOfColumnKZ)

                    currentRow += 1
                    index += 1
                    continue

                if "Код" in currentLine:
                    if checkToKZ:
                        self.fillCell("средн.", 3 + sdvigY, startOffset + columnOffset + 1 + sdvigX + 1)
                        self.fillCell("макс.", 3 + sdvigY, startOffset + columnOffset + 1 + sdvigX)
                        self.sheet.merge_cells(start_row=2 + sdvigY, end_row=2 + sdvigY,
                                               start_column=startOffset + columnOffset + 1 + sdvigX,
                                               end_column=startOffset + columnOffset + 1 + sdvigX + 1)
                        self.fillCell("постоянная часть", 2 + sdvigY, startOffset + columnOffset + sdvigX + 1)

                        self.fillCell("эквив.", 3 + sdvigY, startOffset + columnOffset + 2 + sdvigX + 1)
                        self.fillCell("макс.", 3 + sdvigY, startOffset + columnOffset + 3 + sdvigX + 1)
                        self.sheet.merge_cells(start_row=2 + sdvigY, end_row=2 + sdvigY,
                                               start_column=startOffset + columnOffset + 2 + sdvigX + 1,
                                               end_column=startOffset + columnOffset + 3 + sdvigX + 1)
                        self.fillCell("полуразмах", 2 + sdvigY, startOffset + columnOffset + 2 + sdvigX + 1)
                        self.drawAllInColumn(2 + sdvigY, startOffset + columnOffset + 2 + sdvigX + 1, startOffset + columnOffset + 3 + sdvigX + 1)
                    else:
                        self.sheet.merge_cells(start_row=2 + sdvigY, end_row=3 + sdvigY,
                                               start_column=startOffset + columnOffset + 1 + sdvigX,
                                               end_column=startOffset + columnOffset + 1 + sdvigX)
                        self.fillCell("стат.", 2 + sdvigY, startOffset + columnOffset + 1 + sdvigX)
                        self.sheet.merge_cells(start_row=2 + sdvigY, end_row=2 + sdvigY,
                                               start_column=startOffset + columnOffset + 2 + sdvigX,
                                               end_column=startOffset + columnOffset + 3 + sdvigX)
                        self.fillCell("полуразмах", 2 + sdvigY, startOffset + columnOffset + 2 + sdvigX)
                        self.drawAllInColumn(2 + sdvigY, startOffset + columnOffset + 2 + sdvigX, startOffset + columnOffset + 3 + sdvigX)
                        self.fillCell("эквив.", 3 + sdvigY, startOffset + columnOffset + 2 + sdvigX)
                        self.fillCell("макс.", 3 + sdvigY, startOffset + columnOffset + 3 + sdvigX)

                    index += 1

                else:
                    if checkToMAX:
                        try:
                            maximumOfColumn1 = maximumOfColumn1 if abs(maximumOfColumn1) > abs(float(splittedLine[4].replace(",", ".").replace(u'\xa0', ""))) else float(splittedLine[4].replace(",", ".").replace(u'\xa0', ""))
                        except:
                            maximumOfColumn1 = 0
                        try:
                            maximumOfColumn2 = maximumOfColumn2 if abs(maximumOfColumn2) > abs(float(splittedLine[8].replace(",", ".").replace(u'\xa0', ""))) else float(splittedLine[8].replace(",", ".").replace(u'\xa0', ""))
                        except:
                            maximumOfColumn2 = 0
                        try:
                            maximumOfColumn3 = maximumOfColumn3 if abs(maximumOfColumn3) > abs(float(splittedLine[12].replace(",", ".").replace(u'\xa0', ""))) else float(splittedLine[12].replace(",", ".").replace(u'\xa0', ""))
                        except:
                            maximumOfColumn3 = 0
                        # maximumOfColumn2 = max(maximumOfColumn2, float(splittedLine[8].replace(" ", "").replace(u'\xa0', "").replace(",", ".")))
                        # maximumOfColumn3 = max(maximumOfColumn3, float(splittedLine[12].replace(" ", "").replace(u'\xa0', "").replace(",", ".")))

                    if flagToWriteNames == False:
                        self.fillCell(splittedLine[0], currentRow + 1 + sdvigY, 3)

                    if checkToKZ:
                        # 4 -> 1 (KZ)
                        try:
                            KZValue = splittedLine[13].replace(",", ".").replace(u'\xa0', "") if abs(
                                float(splittedLine[13].replace(",", ".").replace(u'\xa0', ""))) > abs(
                                float(splittedLine[15].replace(",", ".").replace(u'\xa0', ""))) else \
                                splittedLine[15].replace(",", ".").replace(u'\xa0', "")
                            if checkToMAX:
                                maximumOfColumnKZ = maximumOfColumnKZ if abs(maximumOfColumnKZ) > abs(
                                    float(KZValue)) else float(KZValue)
                            self.fillCell(KZValue, currentRow + 1 + sdvigY, startOffset + columnOffset + 1 + sdvigX)
                        except:
                            self.fillCell("X", currentRow + 1 + sdvigY, startOffset + columnOffset + 1 + sdvigX)

                        # 1 -> 2
                        try:
                            self.fillCell(splittedLine[4].replace(",", "."), currentRow + 1 + sdvigY,
                                          startOffset + columnOffset + 2 + sdvigX)
                        except:
                            self.fillCell("X", currentRow + 1 + sdvigY, startOffset + columnOffset + 2 + sdvigX)

                        # 2 -> 3
                        try:
                            self.fillCell(splittedLine[8].replace(",", "."), currentRow + 1 + sdvigY,
                                          startOffset + columnOffset + 3 + sdvigX)
                        except:
                            self.fillCell("X", currentRow + 1 + sdvigY, startOffset + columnOffset + 3 + sdvigX)

                        # 3 -> 4
                        try:
                            self.fillCell(splittedLine[12].replace(",", "."), currentRow + 1 + sdvigY,
                                          startOffset + columnOffset + 4 + sdvigX)
                        except:
                            self.fillCell("X", currentRow + 1 + sdvigY, startOffset + columnOffset + 4 + sdvigX)

                    else:
                        try:
                            self.fillCell(splittedLine[4].replace(",", "."), currentRow + 1 + sdvigY, startOffset + columnOffset + 1 + sdvigX)
                        except:
                            self.fillCell("X", currentRow + 1 + sdvigY, startOffset + columnOffset + 1 + sdvigX)
                        try:
                            self.fillCell(splittedLine[8].replace(",", "."), currentRow + 1 + sdvigY, startOffset + columnOffset + 2 + sdvigX)
                        except:
                            self.fillCell("X", currentRow + 1 + sdvigY, startOffset + columnOffset + 2 + sdvigX)
                        try:
                            self.fillCell(splittedLine[12].replace(",", "."), currentRow + 1 + sdvigY, startOffset + columnOffset + 3 + sdvigX)
                        except:
                            self.fillCell("X", currentRow + 1 + sdvigY, startOffset + columnOffset + 3 + sdvigX)

                        # KZ logic not needed here as checkToKZ is false

            currentRow += 1

            index += 1
        self.fiillingLeftColumn(numberOfLines, sdvigY, lowerSdvigX)
        self.sheet.merge_cells(start_row=4 + sdvigY, end_row=numberOfLines + sdvigY - 1, start_column=lowerSdvigX,
                                   end_column=lowerSdvigX)

        self.workWithMax(numberOfLines, startOffset, columnOffset + sdvigX, checkToKZ, maximumOfColumn1,
                         maximumOfColumn2, maximumOfColumn3, maximumOfColumnKZ)

        workbook.save(filenameToSave)
        workbook.close()
        self.gifLabel1.hide()
        self.gifLabel1.hide()
        self.gifLabel1.hide()
        self.gifLabel1.hide()
        self.gifLabel1.hide()
        self.gifLabel1.hide()
        self.gifLabel1.hide()
        self.gifLabel1.hide()
        self.gifLabel1.hide()
        self.gifLabel1.hide()
        self.gifLabel1.hide()
        self.gifLabel1.hide()
        self.gifLabel1.hide()
        self.gifLabel1.hide()
        self.gifLabel1.hide()
        self.gifLabel1.hide()
        self.gifLabel1.hide()
        self.gifLabel1.hide()
        self.gifLabel1.hide()
        self.ErrorLabel.setText(f"Готово")
        end_time = datetime.datetime.now()
        print(f"Программа выполнялась {(end_time - start_time).seconds} секунд")

    def buttonPressed(self):
        self.ErrorLabel.setText("")

        msg_box = QMessageBox()
        icon_path = os.path.join(self.getIconPath(), "myicon.ico")
        msg_box.setWindowIcon(QtGui.QIcon(icon_path))
        msg_box.setWindowTitle("Подтверждение")
        msg_box.setText("Помни!!! Только ты несёшь ответственность за достоверность данных.      ")
        msg_box.setStandardButtons(QMessageBox.Ok | QMessageBox.Cancel)

        result = msg_box.exec_()

        if result == QMessageBox.Ok:
            options = QtWidgets.QFileDialog.Options()
            self.pathToSave, _ = QtWidgets.QFileDialog.getSaveFileName(MainWindow, "Сохранить файл", "",
                                                                       "All Files (*);;", options=options)
            self.pathToSave = self.pathToSave + ".xlsx"

            if self.pathToSave and self.filePath != "":
                self.makingsheetWithRules(self.filePath, self.pathToSave)
            else:
                self.gifLabel1.hide()
                self.gifLabel2.hide()
                self.ErrorLabel.setText("Введите файл для обработки")
        else:
            pass

    def buttonPressedGetBD(self):
        self.ErrorLabel.setText("")

        options = QtWidgets.QFileDialog.Options()
        self.pathToSave, _ = QtWidgets.QFileDialog.getSaveFileName(MainWindow, "Сохранить файл", "",
                                                                   "All Files (*);;", options=options)
        self.pathToSave = self.pathToSave + ".xlsx"

        if self.pathToSave and self.filePath != "":
            self.makingBD()
        else:
            self.gifLabel1.hide()
            self.ErrorLabel.setText("Введите файл для обработки")

    def buttonPressedExpressAnal(self):
        self.ErrorLabel.setText("")

        msg_box = QMessageBox()
        icon_path = os.path.join(self.getIconPath(), "myicon.ico")
        msg_box.setWindowIcon(QtGui.QIcon(icon_path))
        msg_box.setWindowTitle("Подтверждение")
        msg_box.setText("Помни!!! Только ты несёшь ответственность за достоверность данных.      ")
        msg_box.setStandardButtons(QMessageBox.Ok | QMessageBox.Cancel)

        result = msg_box.exec_()

        options = QtWidgets.QFileDialog.Options()
        self.fileBD, _ = QtWidgets.QFileDialog.getOpenFileName(MainWindow, "Выберите файл", "", "All Files (*)",
                                                                 options=options)

        if ".xlsx" in self.fileBD:
            self.ErrorLabel.setText("")
        else:
            self.DisplayPath.setText("Неверный формат файла")
            return

        if result == QMessageBox.Ok:
            options = QtWidgets.QFileDialog.Options()
            self.pathToSave, _ = QtWidgets.QFileDialog.getSaveFileName(MainWindow, "Сохранить файл", "",
                                                                       "All Files (*);;", options=options)
            self.pathToSave = self.pathToSave + ".xlsx"

            if self.pathToSave and self.filePath != "":
                self.makingExpressAnal()
            else:
                self.gifLabel1.hide()
                self.ErrorLabel.setText("Введите файл для обработки")
        else:
            pass

    def makingBD(self):
        workbook = openpyxl.Workbook()
        self.sheet = workbook.active

        with open(self.filePath, 'r', encoding='cp1251') as file:
            lines = file.readlines()

        currentRow = 4

        widthOfParam = len("Параметр")
        index = 0

        self.fillCell("Параметр", 2, 1)
        self.sheet.merge_cells(start_row=2, end_row=3, start_column=1, end_column=1)
        self.fillCell("Уведомительное", 2, 2)
        self.sheet.merge_cells(start_row=2, end_row=2, start_column=2, end_column=3)
        self.fillCell("Предельное", 2, 4)
        self.sheet.merge_cells(start_row=2, end_row=2, start_column=4, end_column=5)
        self.fillCell("Пост.", 3, 2)
        self.fillCell("Пер.", 3, 3)
        self.fillCell("Пост.", 3, 4)
        self.fillCell("Пер.", 3, 5)


        while index < len(lines):
            currentLine = lines[index].strip()
            splittedLine = currentLine.split("	")

            if "Параметр:" in currentLine:
                widthOfParam = max(widthOfParam, len(splittedLine[0][10:]))
                self.fillCell(splittedLine[0][10:], currentRow, 1)

                currentRow += 1

            index += 1

        self.sheet.column_dimensions["A"].width = widthOfParam + int(widthOfParam * 0.5)

        workbook.save(self.pathToSave)
        workbook.close()
        self.gifLabel1.hide()
        self.ErrorLabel.setText(f"Готово")

    def drawingForAnal(self, startRow, endRow, value1, value2, column):
        yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
        gray_fill = PatternFill(start_color='D3D3D3', end_color='D3D3D3', fill_type='solid')
        red_font = Font(name="Times New Roman", size=12, color="FF0000", bold=True)
        # bold_font = Font(name="Times New Roman", size=12, bold=True)

        for i in range(startRow, endRow + 1):
            if abs(self.sheet.cell(row=i, column=column).value) >= abs(value1) and abs(self.sheet.cell(row=i, column=column).value) < abs(value2):
                # self.sheet.cell(row=i, column=column).font = bold_font
                self.sheet.cell(row=i, column=column).fill = gray_fill
            elif abs(self.sheet.cell(row=i, column=column).value) >= abs(value2):
                self.sheet.cell(row=i, column=column).font = red_font
                self.sheet.cell(row=i, column=column).fill = yellow_fill

            if abs(self.sheet.cell(row=i, column=column + 1).value) >= abs(value1) and abs(self.sheet.cell(row=i, column=column + 1).value) < abs(value2):
                # self.sheet.cell(row=i, column=column + 1).font = bold_font
                self.sheet.cell(row=i, column=column + 1).fill = gray_fill
            elif abs(self.sheet.cell(row=i, column=column + 1).value) >= abs(value2):
                self.sheet.cell(row=i, column=column + 1).font = red_font
                self.sheet.cell(row=i, column=column + 1).fill = yellow_fill

    def makingExpressAnal(self):
        start_time = datetime.datetime.now()

        self.makingsheetExpressAnal()

        workbook = load_workbook(self.pathToSave)
        self.sheet = workbook.active
        workbookBD = load_workbook(self.fileBD)
        self.sheetBD = workbookBD.active

        lenOfColumn = 0
        for cell in self.sheet['C']:
            if cell.value:
                lenOfColumn += 1

        currColumn = 7

        # param = self.sheet.cell(row=2, column=currColumn).value.replace(", кгс/мм²", "") if self.sheet.cell(row=2, column=currColumn).value is not None else self.sheet.cell(row=2, column=currColumn).value
        param = self.sheet.cell(row=2, column=currColumn).value.split(",")[0] if self.sheet.cell(row=2, column=currColumn).value is not None else self.sheet.cell(row=2, column=currColumn).value

        self.fillCell("Макс.", lenOfColumn + 5, 6)
        self.fillCell("Мин.", lenOfColumn + 6, 6)
        self.fillCell("", lenOfColumn + 8, 5)
        self.fillCell("", lenOfColumn + 8, 6)
        self.sheet.merge_cells(start_row=lenOfColumn + 8, end_row=lenOfColumn + 8, start_column=4, end_column=6)
        self.fillCell("Ограничение уведомительное", lenOfColumn + 8, 4)
        self.fillCell("", lenOfColumn + 9, 5)
        self.fillCell("", lenOfColumn + 9, 6)
        self.sheet.merge_cells(start_row=lenOfColumn + 9, end_row=lenOfColumn + 9, start_column=4, end_column=6)
        self.fillCell("Ограничение предельное", lenOfColumn + 9, 4)

        yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
        gray_fill = PatternFill(start_color='D3D3D3', end_color='D3D3D3', fill_type='solid')
        red_font = Font(name="Times New Roman", size=12, color="FF0000", bold=True)
        purple_fill = PatternFill(start_color="cc99cc", end_color="cc99cc", fill_type="solid")
        # bold_font = Font(name="Times New Roman", size=12, bold=True)

        wasDone = False

        while param is not None:
            currRowFromBD = 4
            paramFromBD = self.sheetBD.cell(row=currRowFromBD, column=1).value
            while paramFromBD is not None:
                if paramFromBD == param:
                    wasDone = True
                    uvedPost = self.sheetBD.cell(row=currRowFromBD, column=2).value
                    uvedPer = self.sheetBD.cell(row=currRowFromBD, column=3).value
                    predPost = self.sheetBD.cell(row=currRowFromBD, column=4).value
                    predPer = self.sheetBD.cell(row=currRowFromBD, column=5).value

                    if isinstance(uvedPost, (int, float)) and isinstance(predPost, (int, float)):
                        self.fillCell("|" + str(uvedPost) + "|", lenOfColumn + 8, currColumn)
                    else:
                        self.fillCell(uvedPost, lenOfColumn + 8, currColumn)

                    self.sheet.cell(row=lenOfColumn + 8, column=currColumn).fill = gray_fill
                    # self.sheet.cell(row=lenOfColumn + 8, column=currColumn).font = bold_font
                    self.sheet.merge_cells(start_row=lenOfColumn + 8, end_row=lenOfColumn + 8, start_column=currColumn,
                                           end_column=currColumn + 1)

                    if isinstance(uvedPost, (int, float)) and isinstance(predPost, (int, float)):
                        self.fillCell("|" + str(predPost) + "|", lenOfColumn + 9, currColumn)
                    else:
                        self.fillCell(predPost, lenOfColumn + 9, currColumn)

                    self.sheet.cell(row=lenOfColumn + 9, column=currColumn).fill = yellow_fill
                    self.sheet.cell(row=lenOfColumn + 9, column=currColumn).font = red_font
                    self.sheet.merge_cells(start_row=lenOfColumn + 9, end_row=lenOfColumn + 9, start_column=currColumn,
                                           end_column=currColumn + 1)

                    if isinstance(uvedPost, (int, float)) and isinstance(predPost, (int, float)):
                        self.drawingForAnal(5, lenOfColumn + 3, uvedPost, predPost, currColumn)

                    self.fillCell(uvedPer, lenOfColumn + 8, currColumn + 2)
                    self.sheet.cell(row=lenOfColumn + 8, column=currColumn + 2).fill = gray_fill
                    # self.sheet.cell(row=lenOfColumn + 8, column=currColumn + 2).font = bold_font
                    self.sheet.merge_cells(start_row=lenOfColumn + 8, end_row=lenOfColumn + 8,
                                           start_column=currColumn + 2, end_column=currColumn + 3)

                    self.fillCell(predPer, lenOfColumn + 9, currColumn + 2)
                    self.sheet.cell(row=lenOfColumn + 9, column=currColumn + 2).fill = yellow_fill
                    self.sheet.cell(row=lenOfColumn + 9, column=currColumn + 2).font = red_font
                    self.sheet.merge_cells(start_row=lenOfColumn + 9, end_row=lenOfColumn + 9,
                                           start_column=currColumn + 2, end_column=currColumn + 3)

                    if isinstance(uvedPer, (int, float)) and isinstance(predPer, (int, float)):
                        self.drawingForAnal(5, lenOfColumn + 3, uvedPer, predPer, currColumn + 2)

                    break

                currRowFromBD += 1
                paramFromBD = self.sheetBD.cell(row=currRowFromBD, column=1).value

            if wasDone is False:
                self.fillCell("", lenOfColumn + 8, currColumn)
                self.sheet.cell(row=lenOfColumn + 8, column=currColumn).fill = purple_fill
                self.sheet.merge_cells(start_row=lenOfColumn + 8, end_row=lenOfColumn + 8, start_column=currColumn,
                                       end_column=currColumn + 1)

                self.fillCell("", lenOfColumn + 9, currColumn)
                self.sheet.cell(row=lenOfColumn + 9, column=currColumn).fill = purple_fill
                self.sheet.merge_cells(start_row=lenOfColumn + 9, end_row=lenOfColumn + 9, start_column=currColumn,
                                       end_column=currColumn + 1)

                self.fillCell("", lenOfColumn + 8, currColumn + 2)
                self.sheet.cell(row=lenOfColumn + 8, column=currColumn + 2).fill = purple_fill
                self.sheet.merge_cells(start_row=lenOfColumn + 8, end_row=lenOfColumn + 8,
                                       start_column=currColumn + 2, end_column=currColumn + 3)

                self.fillCell("", lenOfColumn + 9, currColumn + 2)
                self.sheet.cell(row=lenOfColumn + 9, column=currColumn + 2).fill = purple_fill
                self.sheet.merge_cells(start_row=lenOfColumn + 9, end_row=lenOfColumn + 9,
                                       start_column=currColumn + 2, end_column=currColumn + 3)
            wasDone = False
            currColumn += 4
            param = self.sheet.cell(row=2, column=currColumn).value.split(",")[0] if self.sheet.cell(row=2, column=currColumn).value is not None else self.sheet.cell(row=2, column=currColumn).value

        workbook.save(self.pathToSave)
        workbook.close()
        workbookBD.close()
        self.gifLabel1.hide()
        self.ErrorLabel.setText(f"Готово")
        end_time = datetime.datetime.now()
        print(f"Программа выполнялась {(end_time - start_time).seconds} секунд")

    def makingsheetExpressAnal(self):
        filenameToSave = self.pathToSave
        path = self.filePath
        checkToKZ = 1
        checkToSI = True  # Check to place SI
        checkToMAX = 1
        BLOCK_SIZE = 3 if checkToKZ == 0 else 4
        sdvigX = 4
        sdvigY = 1
        lowerSdvigX = 2
        sdvigToKZ = 1 if checkToKZ else 0
        workbook = openpyxl.Workbook()
        self.sheet = workbook.active
        startOffset = 2
        flagToDrawStart = True
        flagToWriteNames = False

        maximumOfColumn1 = -10000000000000; maximumOfColumn2 = -10000000000000; maximumOfColumn3 = -10000000000000; maximumOfColumnKZ = -10000000000000
        minimumOfColumn1 = 10000000000000; minimumOfColumn2 = 10000000000000; minimumOfColumn3 = 10000000000000; minimumOfColumnKZ = 10000000000000

        with open(path, 'r', encoding='cp1251') as file:
            lines = file.readlines()

        currentRow = 1
        numberOfLines = 0
        columnOffset = -BLOCK_SIZE

        flag = False
        index = 0

        self.fillCell("Ид-р реж.", 1 + sdvigY, 1 + lowerSdvigX)
        self.fillCell("Наименование режима", 1 + sdvigY, 2 + lowerSdvigX)
        self.fillCell("Vпр, км/ч", 1 + sdvigY, 3 + lowerSdvigX)
        self.fillCell("Hабс, м", 1 + sdvigY, 4 + lowerSdvigX)

        self.sheet.merge_cells(start_row=1 + sdvigY, end_row=3 + sdvigY, start_column=lowerSdvigX,
                               end_column=lowerSdvigX)
        self.sheet.merge_cells(start_row=1 + sdvigY, end_row=3 + sdvigY, start_column=1 + lowerSdvigX,
                               end_column=1 + lowerSdvigX)
        self.sheet.merge_cells(start_row=1 + sdvigY, end_row=3 + sdvigY, start_column=2 + lowerSdvigX,
                               end_column=2 + lowerSdvigX)
        self.sheet.merge_cells(start_row=1 + sdvigY, end_row=3 + sdvigY, start_column=3 + lowerSdvigX,
                               end_column=3 + lowerSdvigX)
        self.sheet.merge_cells(start_row=1 + sdvigY, end_row=3 + sdvigY, start_column=4 + lowerSdvigX,
                               end_column=4 + lowerSdvigX)

        self.sheet.column_dimensions["C"].width = len("Ид-р реж.") + 4
        self.sheet.column_dimensions["D"].width = len("Наименование режима") + 5
        self.sheet.column_dimensions["E"].width = len("Vпр, км/ч") + 4
        self.sheet.column_dimensions["F"].width = len("Hабс, м") + 4

        while index < len(lines):

            currentLine = lines[index].strip()
            splittedLine = currentLine.split("	")

            if "Результаты расчета" in currentLine:
                flag = True
                index += 1
                continue

            if "Параметр:" in currentLine:
                maximumOfColumn1 = -10000000000000; maximumOfColumn2 = -10000000000000; maximumOfColumn3 = -10000000000000; maximumOfColumnKZ = -10000000000000
                minimumOfColumn1 = 10000000000000; minimumOfColumn2 = 10000000000000; minimumOfColumn3 = 10000000000000; minimumOfColumnKZ = 10000000000000

                columnOffset += BLOCK_SIZE

                # To draw starting 2 columns
                if columnOffset >= BLOCK_SIZE and flagToDrawStart == True:
                    flagToDrawStart = False

                currentRow = 1
                if checkToSI:
                    self.fillCell(splittedLine[0][10:] + ", " + splittedLine[3][14:], currentRow + sdvigY,
                                  startOffset + columnOffset + 1 + sdvigX)
                else:
                    self.fillCell(splittedLine[0][10:], currentRow + sdvigY, startOffset + columnOffset + 1 + sdvigX)
                self.sheet.merge_cells(start_row=1 + sdvigY, end_row=1 + sdvigY,
                                       start_column=startOffset + columnOffset + 1 + sdvigX,
                                       end_column=startOffset + columnOffset + BLOCK_SIZE + sdvigX)
                index += 1
                continue

            if flag:
                if "" == splittedLine[0]:
                    if columnOffset == 0 & currentRow != 1:
                        numberOfLines = currentRow + 1
                    if checkToMAX:
                        if numberOfLines != 0:
                            self.workWithMinAndMaxBD(numberOfLines, startOffset, columnOffset + sdvigX, checkToKZ,
                                             maximumOfColumn1,
                                             maximumOfColumn2, maximumOfColumn3, maximumOfColumnKZ, 0)
                            self.workWithMinAndMaxBD(numberOfLines, startOffset, columnOffset + sdvigX, checkToKZ,
                                                     minimumOfColumn1,
                                                     minimumOfColumn2, minimumOfColumn3, minimumOfColumnKZ, 1)

                    currentRow += 1
                    index += 1
                    continue

                if "Код" in currentLine:
                    if checkToKZ:
                        self.fillCell("средн.", 3 + sdvigY, startOffset + columnOffset + 1 + sdvigX + 1)
                        self.fillCell("макс.", 3 + sdvigY, startOffset + columnOffset + 1 + sdvigX)
                        self.sheet.merge_cells(start_row=2 + sdvigY, end_row=2 + sdvigY,
                                               start_column=startOffset + columnOffset + 1 + sdvigX,
                                               end_column=startOffset + columnOffset + 1 + sdvigX + 1)
                        self.fillCell("постоянная часть", 2 + sdvigY, startOffset + columnOffset + sdvigX + 1)

                        self.fillCell("эквив.", 3 + sdvigY, startOffset + columnOffset + 2 + sdvigX + 1)
                        self.fillCell("макс.", 3 + sdvigY, startOffset + columnOffset + 3 + sdvigX + 1)
                        self.sheet.merge_cells(start_row=2 + sdvigY, end_row=2 + sdvigY,
                                               start_column=startOffset + columnOffset + 2 + sdvigX + 1,
                                               end_column=startOffset + columnOffset + 3 + sdvigX + 1)
                        self.fillCell("полуразмах", 2 + sdvigY, startOffset + columnOffset + 2 + sdvigX + 1)
                        self.drawAllInColumn(2 + sdvigY, startOffset + columnOffset + 2 + sdvigX + 1,
                                             startOffset + columnOffset + 3 + sdvigX + 1)
                    else:
                        self.sheet.merge_cells(start_row=2 + sdvigY, end_row=3 + sdvigY,
                                               start_column=startOffset + columnOffset + 1 + sdvigX,
                                               end_column=startOffset + columnOffset + 1 + sdvigX)
                        self.fillCell("стат.", 2 + sdvigY, startOffset + columnOffset + 1 + sdvigX)
                        self.sheet.merge_cells(start_row=2 + sdvigY, end_row=2 + sdvigY,
                                               start_column=startOffset + columnOffset + 2 + sdvigX,
                                               end_column=startOffset + columnOffset + 3 + sdvigX)
                        self.fillCell("полуразмах", 2 + sdvigY, startOffset + columnOffset + 2 + sdvigX)
                        self.drawAllInColumn(2 + sdvigY, startOffset + columnOffset + 2 + sdvigX,
                                             startOffset + columnOffset + 3 + sdvigX)
                        self.fillCell("эквив.", 3 + sdvigY, startOffset + columnOffset + 2 + sdvigX)
                        self.fillCell("макс.", 3 + sdvigY, startOffset + columnOffset + 3 + sdvigX)

                    index += 1

                else:
                    if checkToMAX:
                        try:
                            minimumOfColumn1 = minimumOfColumn1 if minimumOfColumn1 < float(splittedLine[4].replace(",", ".").replace(u'\xa0', "")) else float(splittedLine[4].replace(",", ".").replace(u'\xa0', ""))
                        except:
                            minimumOfColumn1 = 0
                        try:
                            minimumOfColumn2 = min(minimumOfColumn2, float(
                            splittedLine[8].replace(" ", "").replace(u'\xa0', "").replace(",", ".")))
                        except:
                            minimumOfColumn2 = 0
                        try:
                            minimumOfColumn3 = min(minimumOfColumn3, float(
                            splittedLine[12].replace(" ", "").replace(u'\xa0', "").replace(",", ".")))
                        except:
                            minimumOfColumn3 = 0

                        try:
                            maximumOfColumn1 = maximumOfColumn1 if maximumOfColumn1 > float(splittedLine[4].replace(",", ".").replace(u'\xa0', "")) else float(splittedLine[4].replace(",", ".").replace(u'\xa0', ""))
                        except:
                            maximumOfColumn1 = 0
                        try:
                            maximumOfColumn2 = max(maximumOfColumn2, float(
                            splittedLine[8].replace(" ", "").replace(u'\xa0', "").replace(",", ".")))
                        except:
                            maximumOfColumn2 = 0
                        try:
                            maximumOfColumn3 = max(maximumOfColumn3, float(
                            splittedLine[12].replace(" ", "").replace(u'\xa0', "").replace(",", ".")))
                        except:
                            maximumOfColumn3 = 0

                    if flagToWriteNames == False:
                        self.fillCell(splittedLine[0], currentRow + 1 + sdvigY, 3)

                    if checkToKZ:
                        try:
                            KZValue = splittedLine[13].replace(",", ".").replace(u'\xa0', "") if abs(
                                float(splittedLine[13].replace(",", ".").replace(u'\xa0', ""))) > abs(
                                float(splittedLine[15].replace(",", ".").replace(u'\xa0', ""))) else \
                                splittedLine[15].replace(",", ".").replace(u'\xa0', "")
                            if checkToMAX:
                                maximumOfColumnKZ = maximumOfColumnKZ if maximumOfColumnKZ > float(KZValue) else float(KZValue)
                                minimumOfColumnKZ = minimumOfColumnKZ if minimumOfColumnKZ < float(KZValue) else float(KZValue)
                            self.fillCell(KZValue, currentRow + 1 + sdvigY, startOffset + columnOffset + 1 + sdvigX)
                        except:
                            self.fillCell("X", currentRow + 1 + sdvigY, startOffset + columnOffset + 1 + sdvigX)

                        try:
                            self.fillCell(splittedLine[4].replace(",", "."), currentRow + 1 + sdvigY,
                                          startOffset + columnOffset + 2 + sdvigX)
                        except:
                            self.fillCell("X", currentRow + 1 + sdvigY,
                                          startOffset + columnOffset + 2 + sdvigX)
                        try:
                            self.fillCell(splittedLine[8].replace(",", "."), currentRow + 1 + sdvigY,
                                          startOffset + columnOffset + 3 + sdvigX)
                        except:
                            self.fillCell("X", currentRow + 1 + sdvigY,
                                          startOffset + columnOffset + 3 + sdvigX)
                        try:
                            self.fillCell(splittedLine[12].replace(",", "."), currentRow + 1 + sdvigY,
                                          startOffset + columnOffset + 4 + sdvigX)
                        except:
                            self.fillCell("X", currentRow + 1 + sdvigY,
                                          startOffset + columnOffset + 4 + sdvigX)

                    else:
                        try:
                            self.fillCell(splittedLine[4].replace(",", "."), currentRow + 1 + sdvigY,
                                          startOffset + columnOffset + 1 + sdvigX)
                        except:
                            self.fillCell("X", currentRow + 1 + sdvigY,
                                          startOffset + columnOffset + 1 + sdvigX)
                        try:
                            self.fillCell(splittedLine[8].replace(",", "."), currentRow + 1 + sdvigY,
                                          startOffset + columnOffset + 2 + sdvigX)
                        except:
                            self.fillCell("X", currentRow + 1 + sdvigY,
                                          startOffset + columnOffset + 2 + sdvigX)
                        try:
                            self.fillCell(splittedLine[12].replace(",", "."), currentRow + 1 + sdvigY,
                                          startOffset + columnOffset + 3 + sdvigX)
                        except:
                            self.fillCell("X", currentRow + 1 + sdvigY,
                                          startOffset + columnOffset + 3 + sdvigX)

            currentRow += 1

            index += 1
        self.fiillingLeftColumn(numberOfLines, sdvigY, lowerSdvigX)
        self.sheet.merge_cells(start_row=4 + sdvigY, end_row=numberOfLines + sdvigY - 1, start_column=lowerSdvigX,
                               end_column=lowerSdvigX)

        self.workWithMinAndMaxBD(numberOfLines, startOffset, columnOffset + sdvigX, checkToKZ, maximumOfColumn1,
                         maximumOfColumn2, maximumOfColumn3, maximumOfColumnKZ, 0)
        self.workWithMinAndMaxBD(numberOfLines, startOffset, columnOffset + sdvigX, checkToKZ, minimumOfColumn1,
                         minimumOfColumn2, minimumOfColumn3, minimumOfColumnKZ, 1)

        workbook.save(filenameToSave)
        workbook.close()

    def workWithMinAndMaxBD(self, numberOfLines, startOffset, columnOffset, checkToKZ, maximumOfColumn1, maximumOfColumn2,
                    maximumOfColumn3, maximumOfColumnKZ, min):
        sdvigY = 1

        target1 = maximumOfColumn1
        target2 = maximumOfColumn2
        target3 = maximumOfColumn3
        targetKZ = maximumOfColumnKZ

        if checkToKZ:
            target1 = maximumOfColumnKZ
            target2 = maximumOfColumn1
            target3 = maximumOfColumn2
            targetKZ = maximumOfColumn3

        if min == 0:
            self.fillCell(str(target1), numberOfLines + 1 + sdvigY, startOffset + columnOffset + 1)
            self.fillCell(str(target2), numberOfLines + 1 + sdvigY, startOffset + columnOffset + 2)
            self.fillCell(str(target3), numberOfLines + 1 + sdvigY, startOffset + columnOffset + 3)
            self.fillCell(str(targetKZ), numberOfLines + 1 + sdvigY, startOffset + columnOffset + 4)
        else:
            self.fillCell(str(target1), numberOfLines + 2 + sdvigY, startOffset + columnOffset + 1)
            self.fillCell(str(target2), numberOfLines + 2 + sdvigY, startOffset + columnOffset + 2)
            self.fillCell(str(target3), numberOfLines + 2 + sdvigY, startOffset + columnOffset + 3)
            self.fillCell(str(targetKZ), numberOfLines + 2 + sdvigY, startOffset + columnOffset + 4)

    def getIconPath(self):
        if getattr(sys, 'frozen', False):
            # Если скомпилировано с PyInstaller
            return sys._MEIPASS
        else:
            # Если запущено как обычный .py-файл
            return "."

    def kvaziClicked(self):
        self.kvaziCounter += 1
        if self.kvaziCounter == 15:
            self.gifLabel1.hide()
            self.ErrorLabel.setText("ЛЮБЛЮ МАКСИКА ЛАДАНОВА 😍😘❤️")

    def buttonPressedVybro(self):
        self.Error2Label.setText(f"")

        msg_box = QMessageBox()
        msg_box.setWindowTitle("Подтверждение")
        icon_path = os.path.join(self.getIconPath(), "myicon.ico")
        msg_box.setWindowIcon(QtGui.QIcon(icon_path))
        msg_box.setText("Помни!!! Только ты несёшь ответственность за достоверность данных.      ")
        msg_box.setStandardButtons(QMessageBox.Ok | QMessageBox.Cancel)
        text_edit = msg_box.findChild(QtWidgets.QTextEdit)

        result = msg_box.exec_()

        if result == QMessageBox.Ok:
            options = QtWidgets.QFileDialog.Options()
            self.pathToSave, _ = QtWidgets.QFileDialog.getSaveFileName(MainWindow, "Сохранить файл", "",
                                                                       "All Files (*);;", options=options)
            self.pathToSave = self.pathToSave + ".xlsx"

            if self.pathToSave and self.filePath != "":
                self.makingSheetForVybro()
            else:
                self.gifLabel2.hide()
                self.Error2Label.setText("Введите файл для обработки")
        else:
            pass

    def drawBigger(self, value, row, column):
        font_style = Font(name="Times New Roman", size=12)
        center_alignment = Alignment(horizontal='center', vertical='center')

        double = Side(border_style="thin", color="000000")
        border_style = Border(left=double, right=double, top=double, bottom=double)

        self.sheet.cell(row=row, column=column, value=value).font = font_style
        self.sheet.cell(row=row, column=column).alignment = center_alignment
        self.sheet.cell(row=row, column=column).border = border_style

    def makingSheetForVybro(self):
        start_time = datetime.datetime.now()
        workbook = openpyxl.Workbook()
        self.sheet = workbook.active

        flagToStart = False

        print(self.detect_encoding(self.filePath))

        with open(self.filePath, 'r', encoding=self.detect_encoding(self.filePath)) as file:
            lines = file.readlines()

        currentRow = 1
        columnOffset = 0
        BLOCK_SIZE = 1
        index = 0
        columns = []
        namesForColumns = []
        names = []
        numberOfCols = 0
        flagForR = False
        valueOfRCurrent = []

        while index < len(lines):
            currentLine = lines[index].strip()
            splittedLine = currentLine.split("	")
            print(splittedLine)

            if "Результаты" in currentLine:
                index += 2
                flagToStart = True
                continue

            if flagToStart:
                if "Параметр:" in currentLine:
                    columnOffset += BLOCK_SIZE + 1
                    names.append(splittedLine)
                    index += 4
                    flagForR = False
                    currentRow = 4

                    for i in range(len(valueOfRCurrent)):
                        for j in range(numberOfCols):

                            self.fillCell(str(columns[i * numberOfCols + j][0]), i + 5, columnOffset + j - numberOfCols * 2)
                            self.fillCell(str(columns[i * numberOfCols + j][1]), i + 5, columnOffset + j - numberOfCols)
                        self.fillCell(float(valueOfRCurrent[i].replace(",", ".")), i + 5, columnOffset - numberOfCols * 2 - 1)
                    if (columnOffset - numberOfCols * 2 - 1 != 1):
                        self.sheet.merge_cells(start_row=3, end_row=4, start_column=columnOffset - numberOfCols * 2 - 1,
                                               end_column=columnOffset - numberOfCols * 2 - 1)
                        self.fillCell("С.К.З сумм.", 3, columnOffset - numberOfCols * 2 - 1)
                        self.sheet.column_dimensions[self.sheet.cell(row=3,
                                                                     column=columnOffset - numberOfCols * 2 - 1).column_letter].width = len(
                            "С.К.З сумм.") + 4

                        self.fiillingLeftColumn(int(len(valueOfRCurrent)) + 1, 4, columnOffset - numberOfCols * 2 - 3 + BLOCK_SIZE + 1)
                        self.sheet.merge_cells(start_row=3, end_row=4, start_column=columnOffset - numberOfCols * 2 - 2,
                                               end_column=columnOffset - numberOfCols * 2 - 2)
                        self.drawAllInRow(columnOffset - numberOfCols * 2 - 2, 3, 4)
                        self.fillCell("Наименование режима", 3, columnOffset - numberOfCols * 2 - 2)

                        self.sheet.column_dimensions[self.sheet.cell(row=3, column=columnOffset - numberOfCols * 2 - 2).column_letter].width = len("Наименование режима") + 4
                        self.sheet.merge_cells(start_row=3, end_row=4, start_column=columnOffset - numberOfCols * 2 - 3,
                                               end_column=columnOffset - numberOfCols * 2 - 3)
                        self.drawAllInRow(columnOffset - numberOfCols * 2 - 3, 3, 4)
                        self.fillCell("Ид-р реж.", 3, columnOffset - numberOfCols * 2 - 3)
                        self.sheet.column_dimensions[self.sheet.cell(row=3,
                                                                     column=columnOffset - numberOfCols * 2 - 3).column_letter].width = len(
                            "Ид-р реж.") + 4

                        self.sheet.merge_cells(start_row=2, end_row=2, start_column=columnOffset - numberOfCols * 2 - 3,
                                               end_column=columnOffset - numberOfCols * 2 - 1 + numberOfCols * 2)
                        self.drawAllInColumn(2, columnOffset - numberOfCols * 2 - 3, columnOffset - numberOfCols * 2 - 1 + numberOfCols * 2)
                        self.fillCell(previousName, 2, columnOffset - numberOfCols * 2 - 3)
                    previousName = splittedLine[0][10:] + ", " + splittedLine[3][14:]
                    columns = []
                    valueOfRCurrent = []
                    continue

                if "Р" in splittedLine[0]:
                    valueOfRCurrent.append(splittedLine[8])
                    if len(namesForColumns) != 0:
                        numberOfCols = len(namesForColumns)
                        BLOCK_SIZE = numberOfCols * 2 + 3

                    if flagForR == False:
                        for i in range(len(namesForColumns)):
                            self.fillCell(namesForColumns[i], 4, columnOffset + i - numberOfCols * 2)
                            self.fillCell(namesForColumns[i], 4, columnOffset + i - numberOfCols)
                        if numberOfCols != 0:
                            self.sheet.merge_cells(start_row=3, end_row=3, start_column=columnOffset - numberOfCols * 2,
                                                  end_column=columnOffset - numberOfCols * 2 + numberOfCols - 1)
                            self.fillCell("С.К.З. в полосах", 3, columnOffset - numberOfCols * 2)
                            self.sheet.merge_cells(start_row=3, end_row=3, start_column=columnOffset - numberOfCols,
                                                  end_column=columnOffset - numberOfCols + numberOfCols - 1)
                            self.drawAllInColumn(3, columnOffset - numberOfCols, columnOffset - numberOfCols + numberOfCols - 1)
                            self.fillCell("Sxx в полосах", 3, columnOffset - numberOfCols)
                        flagForR = True

                    currentRow += 1
                    namesForColumns = []
                    index += 1

                    continue

                if splittedLine[0] != "":
                    columns.append([float(splittedLine[3].replace(",", ".")), "{:.6f}".format(float(splittedLine[4].replace(",", ".")))])
                    firstVal = splittedLine[1].replace(",", ".").split(".")[0]
                    seconVal = splittedLine[2].replace(",", ".").split(".")[0]
                    namesForColumns.append(f"{firstVal}...{seconVal}")

            index += 1

        for i in range(numberOfCols):
            self.fillCell(namesForColumns[i], 4, columnOffset + i - numberOfCols * 2 + BLOCK_SIZE + 1)
            self.fillCell(namesForColumns[i], 4, columnOffset + i - numberOfCols + BLOCK_SIZE + 1)
        self.sheet.merge_cells(start_row=3, end_row=3, start_column=columnOffset - numberOfCols * 2 + BLOCK_SIZE + 1,
                               end_column=columnOffset - numberOfCols * 2 + numberOfCols - 1 + BLOCK_SIZE + 1)
        self.fillCell("С.К.З. в полосах", 3, columnOffset - numberOfCols * 2 + BLOCK_SIZE + 1)
        self.sheet.merge_cells(start_row=3, end_row=3, start_column=columnOffset - numberOfCols + BLOCK_SIZE + 1,
                               end_column=columnOffset - numberOfCols + numberOfCols - 1 + BLOCK_SIZE + 1)
        self.drawAllInColumn(3, columnOffset - numberOfCols + BLOCK_SIZE + 1, columnOffset - numberOfCols + numberOfCols - 1 + BLOCK_SIZE + 1)
        self.fillCell("Sxx в полосах", 3, columnOffset - numberOfCols + BLOCK_SIZE + 1)

        for i in range(len(valueOfRCurrent)):
            for j in range(numberOfCols):
                self.fillCell(str(columns[i * numberOfCols + j][0]), i + 5, columnOffset + j - numberOfCols * 2 + BLOCK_SIZE + 1)
                self.fillCell(str(columns[i * numberOfCols + j][1]), i + 5, columnOffset + j - numberOfCols + BLOCK_SIZE + 1)
            self.fillCell(float(valueOfRCurrent[i].replace(",", ".")), i + 5, columnOffset - numberOfCols * 2 - 1 + BLOCK_SIZE + 1)

        self.sheet.merge_cells(start_row=3, end_row=4, start_column=columnOffset - numberOfCols * 2 - 1 + BLOCK_SIZE + 1,
                               end_column=columnOffset - numberOfCols * 2 - 1 + BLOCK_SIZE + 1)
        self.fillCell("С.К.З сумм.", 3, columnOffset - numberOfCols * 2 - 1 + BLOCK_SIZE + 1)
        self.sheet.column_dimensions[self.sheet.cell(row=3,
                                                     column=columnOffset - numberOfCols * 2 - 1 + BLOCK_SIZE + 1).column_letter].width = len(
            "С.К.З сумм.") + 4
        self.fiillingLeftColumn(int(len(valueOfRCurrent)) + 1, 4, columnOffset - numberOfCols * 2 - 3 + BLOCK_SIZE + 1)
        self.sheet.merge_cells(start_row=3, end_row=4,
                               start_column=columnOffset - numberOfCols * 2 - 2 + BLOCK_SIZE + 1,
                               end_column=columnOffset - numberOfCols * 2 - 2 + BLOCK_SIZE + 1)
        self.drawAllInRow(columnOffset - numberOfCols * 2 - 2 + BLOCK_SIZE + 1, 3, 4)
        self.fillCell("Наименование режима", 3, columnOffset - numberOfCols * 2 - 2 + BLOCK_SIZE + 1)
        self.sheet.column_dimensions[
            self.sheet.cell(row=3, column=columnOffset - numberOfCols * 2 - 2 + BLOCK_SIZE + 1).column_letter].width = len(
            "Наименование режима") + 4

        self.sheet.merge_cells(start_row=3, end_row=4,
                               start_column=columnOffset - numberOfCols * 2 - 3 + BLOCK_SIZE + 1,
                               end_column=columnOffset - numberOfCols * 2 - 3 + BLOCK_SIZE + 1)
        self.drawAllInRow(columnOffset - numberOfCols * 2 - 3 + BLOCK_SIZE + 1, 3, 4)
        self.fillCell("Ид-р реж.", 3, columnOffset - numberOfCols * 2 - 3 + BLOCK_SIZE + 1)
        self.sheet.column_dimensions[self.sheet.cell(row=3,
                                                     column=columnOffset - numberOfCols * 2 - 3 + BLOCK_SIZE + 1).column_letter].width = len(
            "Ид-р реж.") + 4
        self.sheet.merge_cells(start_row=2, end_row=2, start_column=columnOffset - numberOfCols * 2 - 3 + BLOCK_SIZE + 1,
                               end_column=columnOffset - numberOfCols * 2 - 1 + numberOfCols * 2 + BLOCK_SIZE + 1)
        self.drawAllInColumn(2, columnOffset - numberOfCols * 2 - 3 + BLOCK_SIZE + 1,
                             columnOffset - numberOfCols * 2 - 1 + numberOfCols * 2 + BLOCK_SIZE + 1)
        self.fillCell(previousName, 2, columnOffset - numberOfCols * 2 - 3 + BLOCK_SIZE + 1)

        self.fiillingLeftColumn(len(valueOfRCurrent) + 1, 4, 3)

        workbook.save(self.pathToSave)
        workbook.close()
        self.gifLabel2.hide()
        self.Error2Label.setText(f"Готово")
        end_time = datetime.datetime.now()
        print(f"Программа выполнялась {(end_time - start_time).seconds} секунд")

    def drawAllInColumn(self, row, startCol, endCol):
        double = Side(border_style="thin", color="000000")
        border_style = Border(left=double, right=double, top=double, bottom=double)
        for col in range(startCol, endCol + 1):
            self.sheet.cell(row=row, column=col).border = border_style

    def drawAllInRow(self, column, startRow, endRow):
        double = Side(border_style="thin", color="000000")
        border_style = Border(left=double, right=double, top=double, bottom=double)
        for row in range(startRow, endRow + 1):
            self.sheet.cell(row=row, column=column).border = border_style


# def excepthook(type, value, traceback):
#     sys.__excepthook__(type, value, traceback)


if __name__ == "__main__":
    app = QtWidgets.QApplication(sys.argv)
    MainWindow = QtWidgets.QMainWindow()
    ui = Ui_MainWindow()
    ui.setupUi(MainWindow)
    MainWindow.show()
    sys.exit(app.exec_())

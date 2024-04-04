import openpyxl
from openpyxl.chart import BarChart, Reference
from openpyxl.drawing.text import Paragraph, ParagraphProperties, CharacterProperties, Font
from openpyxl.chart.text import RichText
from openpyxl.chart.layout import Layout, ManualLayout
from openpyxl.drawing.text import RichTextProperties

wb = openpyxl.load_workbook(f'{input("Введите название файла, с которого берутся данные: ")}.xlsx')

sheet = wb.active

numberOfColsInBlock = sheet.cell(row=1, column=1).value
columnOffset = 2

data_x = []
data_y = []

count = 0

font_test = Font(typeface='Times New Roman')
cp = CharacterProperties(latin=font_test, sz=900)
cp1 = CharacterProperties(latin=font_test, sz=1400)

for row in sheet.iter_rows(min_row=3, max_row=sheet.max_row, min_col=1, max_col=1):
    data_x.append(row[0].value)

for i in range(int((sheet.max_column - 1) / numberOfColsInBlock)):
    for col_index in range(0, numberOfColsInBlock):
        data_y = []

        for row in sheet.iter_rows(min_row=3, max_row=sheet.max_row, min_col=col_index + columnOffset, max_col=col_index + columnOffset):
            data_y.append(row[0].value)

        chart = BarChart()
        chart.add_data(Reference(sheet, min_col=col_index + columnOffset, min_row=2, max_row=sheet.max_row), titles_from_data=True)
        chart.set_categories(Reference(sheet, min_col=1, min_row=3, max_row=sheet.max_row))

        chart.title = str(sheet.cell(row=1, column=columnOffset).value) + "\n" + sheet.cell(row=2, column=1).value
        chart.title.tx.rich.p[0].r[0].rPr = cp1
        chart.title.tx.rich.p[1].r[0].rPr = cp1

        chart.x_axis.txPr = RichText(p=[Paragraph(pPr=ParagraphProperties(defRPr=cp), endParaRPr=cp)])
        chart.y_axis.txPr = RichText(p=[Paragraph(pPr=ParagraphProperties(defRPr=cp), endParaRPr=cp)])

        chart.y_axis.title = "g²/Гц" if sheet.cell(row=2, column=columnOffset + col_index).value == "Sxx" else "g"

        chart.y_axis.txPr = chart.y_axis.title.text.rich
        chart.y_axis.txPr.properties.vert = "horz"

        # Move height of Y axis legend
        chart.y_axis.title.layout = Layout(
            manualLayout=ManualLayout(
                h=0.85,  # value between 0 and 1
                x=0,
                y=0.9
            )
        )

        pp = ParagraphProperties(defRPr=cp)
        chart.y_axis.title.tx.rich.p[0].pPr = pp

        sheet.add_chart(chart, f'D1')
        count += 1

    columnOffset += numberOfColsInBlock

# print(count)
# print(sheet.max_row)
# print(int((sheet.max_column - 1) / numberOfColsInBlock))
wb.save('gistogram.xlsx')
